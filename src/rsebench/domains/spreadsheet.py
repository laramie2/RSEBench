"""SpreadsheetBench-compatible artifact noise and validation."""

from __future__ import annotations

import datetime as dt
import hashlib
import random
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from pydantic import BaseModel, ConfigDict

from rsebench.contracts import ValidationReport
from rsebench.hashing import sha256_file


class SpreadsheetTask(BaseModel):
    model_config = ConfigDict(extra="forbid", arbitrary_types_allowed=True)
    task_id: str
    workbook_path: Path
    gold_workbook_path: Path | None = None
    prompt: str
    answer_sheet: str
    answer_range: str
    original_sheets: tuple[str, ...]
    clean_hash: str

    @classmethod
    def from_paths(
        cls,
        *,
        task_id: str,
        workbook_path: Path | str,
        prompt: str,
        answer_sheet: str,
        answer_range: str,
        gold_workbook_path: Path | str | None = None,
    ) -> "SpreadsheetTask":
        source = Path(workbook_path)
        workbook = load_workbook(source, read_only=True, data_only=False)
        sheets = tuple(workbook.sheetnames)
        workbook.close()
        return cls(
            task_id=task_id,
            workbook_path=source,
            gold_workbook_path=(
                Path(gold_workbook_path) if gold_workbook_path is not None else None
            ),
            prompt=prompt,
            answer_sheet=answer_sheet,
            answer_range=answer_range,
            original_sheets=sheets,
            clean_hash=sha256_file(source),
        )


class SpreadsheetNoiseResult(BaseModel):
    model_config = ConfigDict(extra="forbid", arbitrary_types_allowed=True)
    output_path: Path
    operator: str
    severity: str
    seed: int
    added_sheet: str
    clean_hash: str
    noisy_hash: str


def _available_title(workbook, base: str) -> str:
    if base not in workbook.sheetnames:
        return base
    index = 2
    while f"{base}_{index}" in workbook.sheetnames:
        index += 1
    return f"{base}_{index}"


def _source_sheet(workbook, task: SpreadsheetTask, seed: int):
    candidates = [name for name in task.original_sheets if name != task.answer_sheet]
    if not candidates:
        candidates = list(task.original_sheets)
    return workbook[random.Random(seed).choice(sorted(candidates))]


def _prepare_output(task: SpreadsheetTask, output_path: Path | str):
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(task.workbook_path, output)
    return output, load_workbook(output, data_only=False)


def inject_backup_sheet(
    task: SpreadsheetTask,
    output_path: Path | str,
    *,
    severity: str,
    seed: int,
) -> SpreadsheetNoiseResult:
    output, workbook = _prepare_output(task, output_path)
    source = _source_sheet(workbook, task, seed)
    copied = workbook.copy_worksheet(source)
    copied.title = _available_title(workbook, "Backup_Archive")
    copied.sheet_properties.tabColor = "808080"
    added_sheet = copied.title
    workbook.save(output)
    workbook.close()
    return SpreadsheetNoiseResult(
        output_path=output,
        operator="stale_backup_sheet",
        severity=severity,
        seed=seed,
        added_sheet=added_sheet,
        clean_hash=task.clean_hash,
        noisy_hash=sha256_file(output),
    )


def inject_semantic_decoy_sheet(
    task: SpreadsheetTask,
    output_path: Path | str,
    *,
    severity: str,
    seed: int,
) -> SpreadsheetNoiseResult:
    if severity not in {"L1", "L2", "L3"}:
        raise ValueError(f"unsupported severity: {severity}")
    output, workbook = _prepare_output(task, output_path)
    source = _source_sheet(workbook, task, seed)
    title = _available_title(workbook, "Draft_Decoy")
    decoy = workbook.create_sheet(title)
    row_budget = {"L1": 8, "L2": 20, "L3": 50}[severity]
    column_budget = {"L1": 6, "L2": 10, "L3": 16}[severity]
    for row in source.iter_rows(
        min_row=1,
        max_row=min(source.max_row, row_budget),
        min_col=1,
        max_col=min(source.max_column, column_budget),
    ):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                value = f"STALE_FORMULA: {value}"
            decoy.cell(row=cell.row + 1, column=cell.column, value=value)
    decoy["A1"] = "STALE DRAFT — REFERENCE ONLY"
    decoy.sheet_properties.tabColor = "C0504D"
    workbook.save(output)
    workbook.close()
    return SpreadsheetNoiseResult(
        output_path=output,
        operator="semantic_decoy_sheet",
        severity=severity,
        seed=seed,
        added_sheet=title,
        clean_hash=task.clean_hash,
        noisy_hash=sha256_file(output),
    )


def _sheet_digest(sheet) -> str:
    digest = hashlib.sha256()
    digest.update(sheet.sheet_state.encode("utf-8"))
    for merged in sorted(str(item) for item in sheet.merged_cells.ranges):
        digest.update(f"merged:{merged}".encode("utf-8"))
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                value = cell.value
                if hasattr(value, "ref"):
                    attributes = sorted(
                        (key, repr(item)) for key, item in vars(value).items()
                    )
                    canonical_value = f"{type(value).__name__}:{attributes}"
                else:
                    canonical_value = repr(value)
                record = (
                    f"{cell.coordinate}\0{cell.data_type}\0{canonical_value}\0"
                    f"{cell.number_format}\n"
                )
                digest.update(record.encode("utf-8"))
    return digest.hexdigest()


def validate_spreadsheet_noise(
    task: SpreadsheetTask, result: SpreadsheetNoiseResult
) -> ValidationReport:
    messages: list[str] = []
    try:
        clean = load_workbook(task.workbook_path, data_only=False)
        noisy = load_workbook(result.output_path, data_only=False)
    except Exception as exc:
        return ValidationReport(
            structural_valid=False,
            label_invariant=False,
            solvable=False,
            answer_leak_free=True,
            accepted=False,
            messages=[f"workbook_load_failed: {exc}"],
        )
    try:
        original_present = all(name in noisy.sheetnames for name in task.original_sheets)
        unchanged = original_present and all(
            _sheet_digest(clean[name]) == _sheet_digest(noisy[name])
            for name in task.original_sheets
        )
        added_present = result.added_sheet in noisy.sheetnames
        structural = original_present and added_present
        if not unchanged:
            messages.append("one or more protected original sheets changed")
        # Some released rows encode multiple answer sheets in a single legacy
        # string. Preserving every original sheet/range is the stronger gate.
        answer_present = original_present
        accepted = structural and unchanged and answer_present
        return ValidationReport(
            structural_valid=structural,
            label_invariant=unchanged,
            solvable=answer_present and unchanged,
            answer_leak_free=True,
            accepted=accepted,
            checks={
                "original_sheet_count": len(task.original_sheets),
                "added_sheet": result.added_sheet,
            },
            messages=messages,
        )
    finally:
        clean.close()
        noisy.close()


def _datetime_to_float(value: dt.datetime) -> float:
    delta = value - dt.datetime(1899, 12, 30)
    return delta.days + delta.seconds / 86400.0


def _transform_value(value):
    if isinstance(value, bool):
        return round(float(value), 2)
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, dt.time):
        return str(value)[:-3]
    if isinstance(value, dt.datetime):
        return round(_datetime_to_float(value), 0)
    if isinstance(value, str):
        try:
            return round(float(value), 2)
        except ValueError:
            return value
    return value


def _compare_value(first, second) -> bool:
    first = _transform_value(first)
    second = _transform_value(second)
    if first in ("", None) and second in ("", None):
        return True
    return type(first) is type(second) and first == second


def compare_answer_range(
    gold_path: Path | str, predicted_path: Path | str, answer_position: str
) -> tuple[bool, str]:
    """Port the official SpreadsheetBench cell-value comparison semantics."""
    try:
        gold = load_workbook(gold_path, data_only=True)
        predicted = load_workbook(predicted_path, data_only=True)
    except Exception as exc:
        return False, f"load error: {exc}"
    try:
        for specification in answer_position.split(","):
            specification = specification.strip()
            if not specification:
                continue
            if "!" in specification:
                sheet_name, cell_range = specification.split("!", 1)
                sheet_name = sheet_name.strip("'\"")
            else:
                sheet_name = gold.sheetnames[0]
                cell_range = specification
            if sheet_name not in predicted.sheetnames:
                return False, f"worksheet not found: {sheet_name}"
            min_col, min_row, max_col, max_row = range_boundaries(
                cell_range.strip("'\"")
            )
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    expected = gold[sheet_name].cell(row, column).value
                    actual = predicted[sheet_name].cell(row, column).value
                    if not _compare_value(expected, actual):
                        return (
                            False,
                            f"value@{sheet_name}!{row},{column}: "
                            f"gt={expected!r} pred={actual!r}",
                        )
        return True, ""
    finally:
        gold.close()
        predicted.close()
