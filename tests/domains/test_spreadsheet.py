from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.formula import ArrayFormula

from rsebench.domains.spreadsheet import (
    SpreadsheetTask,
    compare_answer_range,
    inject_backup_sheet,
    inject_semantic_decoy_sheet,
    validate_spreadsheet_noise,
)


@pytest.fixture
def spreadsheet_task(tmp_path: Path) -> SpreadsheetTask:
    clean = tmp_path / "clean.xlsx"
    gold = tmp_path / "gold.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["item", "amount"])
    sheet.append(["alpha", 2])
    sheet["C2"] = ArrayFormula(ref="C2:C2", text="=B2")
    answer = workbook.create_sheet("Answer")
    answer["A1"] = "total"
    answer["B1"] = "=SUM(Data!B2:B2)"
    workbook.save(clean)
    answer["B1"] = 2
    workbook.save(gold)
    return SpreadsheetTask.from_paths(
        task_id="sheet-1",
        workbook_path=clean,
        gold_workbook_path=gold,
        prompt="Fill Answer!B1.",
        answer_sheet="Answer",
        answer_range="B1",
    )


def test_backup_sheet_preserves_original_sheets_and_answer(tmp_path, spreadsheet_task):
    result = inject_backup_sheet(
        spreadsheet_task, tmp_path / "noisy.xlsx", severity="L2", seed=42
    )
    workbook = load_workbook(result.output_path, data_only=False)
    assert "Backup_Archive" in workbook.sheetnames
    assert set(spreadsheet_task.original_sheets) <= set(workbook.sheetnames)
    report = validate_spreadsheet_noise(spreadsheet_task, result)
    assert report.structural_valid
    assert report.label_invariant
    assert report.accepted


def test_semantic_decoy_changes_only_new_sheet(tmp_path, spreadsheet_task):
    result = inject_semantic_decoy_sheet(
        spreadsheet_task, tmp_path / "decoy.xlsx", severity="L3", seed=9
    )
    workbook = load_workbook(result.output_path, data_only=False)
    assert result.added_sheet.startswith("Draft_Decoy")
    assert workbook["Data"]["B2"].value == 2
    assert "STALE" in str(workbook[result.added_sheet]["A1"].value)
    assert validate_spreadsheet_noise(spreadsheet_task, result).accepted


def test_answer_range_comparator_matches_official_numeric_rules(spreadsheet_task):
    ok, reason = compare_answer_range(
        spreadsheet_task.gold_workbook_path,
        spreadsheet_task.gold_workbook_path,
        "Answer!B1",
    )
    assert ok, reason
