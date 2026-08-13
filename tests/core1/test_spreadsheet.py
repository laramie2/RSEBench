from __future__ import annotations

from openpyxl import Workbook, load_workbook

from rsebench.core1.spreadsheet import (
    build_spreadsheet_n1_pair,
    build_spreadsheet_n2_pair,
)
from rsebench.domains.spreadsheet import (
    SpreadsheetTask,
    _sheet_digest,
    validate_spreadsheet_noise,
)


def spreadsheet_task(tmp_path) -> SpreadsheetTask:
    workbook = Workbook()
    source = workbook.active
    source.title = "Transactions_2025"
    source.append(["DATE", "REF", "AMOUNT"])
    source.append(["2025-01-01", "A", 100.0])
    source.append(["2025-01-02", "B", 200.0])
    answer = workbook.create_sheet("LISTS")
    answer.append(["DATE", "REF", "AMOUNT"])
    path = tmp_path / "task.xlsx"
    workbook.save(path)
    return SpreadsheetTask.from_paths(
        task_id="sheet-1",
        workbook_path=path,
        prompt=(
            "Combine rows by matching duplicates on the DATE and REF columns, "
            "then sort amounts from lowest to highest and add a TOTAL row."
        ),
        answer_sheet="LISTS",
        answer_range="A2:C4",
    )


def test_n1_appends_one_erroneous_constraint_without_answer_leak(tmp_path) -> None:
    task = spreadsheet_task(tmp_path)

    pair = build_spreadsheet_n1_pair(task, seed=7)

    assert pair.noisy_prompt.startswith(task.prompt)
    assert pair.clean_prompt == task.prompt
    assert pair.changed_axis == "join_key"
    assert pair.original_constraint == "DATE and REF"
    assert pair.misleading_constraint == "DATE only"
    assert pair.noisy_prompt.count("Prior analyst handover") == 1
    assert task.answer_range not in pair.noisy_prompt


def test_n2_adds_unlabeled_prior_sheet_and_preserves_originals(tmp_path) -> None:
    task = spreadsheet_task(tmp_path)
    output = tmp_path / "noisy.xlsx"
    clean = load_workbook(task.workbook_path, data_only=False)
    before = {name: _sheet_digest(clean[name]) for name in clean.sheetnames}
    clean.close()

    result = build_spreadsheet_n2_pair(task, output, seed=7)

    noisy = load_workbook(output, data_only=False)
    try:
        assert len(noisy.sheetnames) == len(task.original_sheets) + 1
        assert result.added_sheet in noisy.sheetnames
        assert all(_sheet_digest(noisy[name]) == digest for name, digest in before.items())
        added_text = " ".join(
            str(cell.value)
            for row in noisy[result.added_sheet].iter_rows()
            for cell in row
            if cell.value is not None
        ).upper()
        for warning in ("STALE", "DECOY", "OLD", "REFERENCE ONLY"):
            assert warning not in added_text
        assert "2024" in added_text
    finally:
        noisy.close()

    report = validate_spreadsheet_noise(task, result)
    assert report.accepted, report.messages
    assert report.label_invariant
